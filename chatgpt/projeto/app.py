from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import os

# Configuração do Flask
app = Flask(__name__)
CORS(app) # Habilita CORS para permitir requisições do frontend

# --- Rotas da Aplicação ---


# Rota para servir a página inicial (index.html)
@app.route('/')
def home():
    # Serve o arquivo index.html do diretório 'static'
    return send_from_directory('static', 'index.html')

# Rota para servir outros arquivos estáticos (CSS, JS, imagens)
@app.route('/<path:path>')
def serve_static(path):
    # Serve qualquer arquivo do diretório 'static'
    return send_from_directory('static', path)

# Rota para receber os dados do formulário via POST
@app.route('/receber-dados', methods=['POST'])
def receber_dados():
    # Obtém os dados JSON da requisição
    data = request.get_json()

    # Extrai os dados dos campos do formulário
    nome = data.get('nome')
    email = data.get('email')
    telefone = data.get('telefone')
    agencia = data.get('agencia') # O nome 'agencia' foi mantido para compatibilidade com o HTML/JS anterior

    # Caminho para o arquivo Excel
    caminho_arquivo = 'leads.xlsx'

    # --- Lógica de escrita no Excel ---

    # Verifica se o arquivo Excel já existe
    if not os.path.exists(caminho_arquivo):
        # Se não existir, cria uma nova pasta de trabalho (workbook)
        wb = Workbook()
        # Seleciona a planilha ativa (geralmente a primeira)
        ws = wb.active
        # Define o título da planilha
        ws.title = "Leads Captados"

        # Define o cabeçalho da planilha
        headers = ['Nome', 'Email', 'Telefone', 'Interesse']

        # --- Estiliza o cabeçalho ---
        # Define a fonte em negrito
        bold_font = Font(bold=True)
        # Define o preenchimento amarelo (use o código HEX para a cor)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Itera sobre o cabeçalho e aplica o estilo
        for col_num, header in enumerate(headers, start=1):
            # Obtém a célula na primeira linha e coluna atual
            cell = ws.cell(row=1, column=col_num)
            # Define o valor da célula como o texto do cabeçalho
            cell.value = header
            # Aplica a fonte em negrito
            cell.font = bold_font
            # Aplica o preenchimento amarelo
            cell.fill = yellow_fill

    else:
        # Se o arquivo existir, carrega a pasta de trabalho existente
        wb = load_workbook(caminho_arquivo)
        # Seleciona a planilha ativa
        ws = wb.active

    # Adiciona os dados do novo lead a uma nova linha na planilha
    ws.append([nome, email, telefone, agencia])

    # Salva as alterações no arquivo Excel
    try:
        wb.save(caminho_arquivo)
        # Retorna uma resposta JSON indicando sucesso
        return jsonify({'status': 'sucesso'})
    except Exception as e:
        # Se houver um erro ao salvar (por exemplo, arquivo aberto), retorne um erro
        print(f"Erro ao salvar arquivo Excel: {e}")
        return jsonify({'status': 'erro', 'mensagem': str(e)}), 500 # Retorna um erro 500

# --- Execução da Aplicação ---

# Executa a aplicação Flask se o script for o principal
if __name__ == '__main__':
    # app.run(debug=True) # Remover debug=True em produção
    app.run(debug=True) # Manter debug=True para desenvolvimento local